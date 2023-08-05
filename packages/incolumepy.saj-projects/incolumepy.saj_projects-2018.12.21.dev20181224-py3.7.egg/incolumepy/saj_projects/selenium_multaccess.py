#!/bin/python
# -*- Coding: utf-8 -*-
# incolumepy.teste.handler_selenium.selenium_multaccess03.incolumepy

import os
import re
import sys
import json
import logging
import logging.handlers
import inspect
import requests
import pandas as pd
from os.path import abspath, basename
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from urllib.error import URLError
from collections import namedtuple
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from incolumepy.utils.files import realfilename
from legis import Legis

Atos = namedtuple('Ato', 'nome numero ano date origem situacao relacionada link_elem_txt elem_txt')


class Scrapy_camara:
    def __init__(self, **kwargs):
        '''

        :param kwargs: driver=str, firefoxbin=str full path, firefox=selenium.webdriver,
        ano=num, numero=num, expressao=str, click=list, collection_query=list,
        collection_metadados = list, collection_atos =[], proxies = dict
        '''
        driver_file = os.path.join(*'drivers/geckodriver'.split(os.sep))
        driver_path = os.path.join(os.path.dirname(__file__), *driver_file.split(os.sep))
        self.driver = abspath(driver_path)
        assert os.path.exists(self.driver), f"{self.driver} falhou."
        # self.driver = '/home/brito/projetos/saj_projects/drivers/geckodriver'
        self.ano = None
        self.numero = None
        self.expressao = None
        self.click = ['apelido', 'Ilei', 'Ileicom', 'Ideclei', 'Idecret']
        self.collection_query = []
        self.collection_metadados = []
        self.collection_metadados_header = []
        self.basename = os.path.basename(__file__).split('.')[0]
        self.proxies = {'http': 'http://10.1.101.101:8080',
                   'https': 'http://10.1.101.101:8080',
                   'ftp': 'http://10.1.101.101:8080',
                   }
        for key, value in kwargs.items():
            self.__dict__[key.lower()] = value
        self.firefoxbin = os.path.abspath(self.driver)
        self.firefox = webdriver.Firefox(executable_path=self.firefoxbin)


    def __del__(self):
        ''' Destructor from firefox instance'''
        self.firefox.close()


    def atos2list(self, container_atos):
        for ato in container_atos:
            d = {}
            d['title'] = ato.text
            d['link'] = ato.find_element_by_tag_name('a').get_attribute('href')
            self.collection_query.append(d)

    def collection_query_toCSV(self, path='', **kwargs):
        '''grava o collection_query em csv'''
        validos = 'list dict'.split()
        dados = ''
        filename = realfilename(os.path.join(path, 'collection_query.csv'))
        for item in validos:
            if item in kwargs:
                dados = kwargs[item]
            else:
                dados = self.collection_query

        df = pd.DataFrame(dados)
        df.to_csv(filename, index=False)
        logging.debug(f"{df}")
        return True

    def collection_query_toJSON(self, path='', **kwargs):
        ''' grava o collection_query em json'''
        validos = 'list dict'.split()
        dados = ''
        filename = realfilename(os.path.join(path, 'collection_query.json'))
        for item in validos:
            if item in kwargs:
                dados = kwargs[item]
            else:
                dados = self.collection_query

        df = pd.DataFrame(dados)
        df.to_json(filename, orient='table')
        logging.debug(f"{df}")
        return True

    def collection_query_toExcel(self, path='', **kwargs):
        '''grava o collection_query em xlsx'''
        validos = 'list dict'.split()
        dados = ''
        filename = realfilename(os.path.join(path, 'collection_query.xlsx'))
        for item in validos:
            if item in kwargs:
                dados = kwargs[item]
            else:
                dados = self.collection_query

        df = pd.DataFrame(dados)
        logging.debug(f"{df}")
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as xslx:
                xslx.book = load_workbook(filename)
                df.to_excel(xslx, index=False)
        except:
            df.to_excel(filename, index=False)

        logging.debug(f"{df}")
        return True

    def collection_metadados_toCSV(self, path='', **kwargs):
        '''grava o collection_metadados em csv'''
        validos = 'list dict'.split()
        dados = ''
        filename = realfilename(os.path.join(path, 'collection_metadados.csv'))
        for item in validos:
            if item in kwargs:
                dados = kwargs[item]
            else:
                dados = self.collection_metadados

        df = pd.DataFrame(dados)
        df.columns = self.collection_metadados_header
        df.to_csv(filename, index=False)
        print(df)
        return True

    def collection_metadados_toJSON(self, path='', **kwargs):
        ''' grava o collection_metadados em json'''
        validos = 'list dict'.split()
        dados = ''
        filename = realfilename(os.path.join(path, 'collection_metadados.json'))
        for item in validos:
            if item in kwargs:
                dados = kwargs[item]
            else:
                dados = self.collection_metadados

        df = pd.DataFrame(dados)
        df.columns = self.collection_metadados_header
        df.to_json(filename, orient='table')
        logging.debug(df)
        return True

    def collection_metadados_toExcel(self, path='', **kwargs):
        '''grava o collection_metadados em xlsx'''
        validos = 'list dict'.split()
        dados = ''
        filename = realfilename(os.path.join(path, 'collection_metadados.xlsx'))
        for item in validos:
            if item in kwargs:
                dados = kwargs[item]
            else:
                dados = self.collection_metadados

        df = pd.DataFrame(dados)
        df.columns = self.collection_metadados_header
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as xslx:
                xslx.book = load_workbook(filename)
                df.to_excel(xslx, index=False)
        except Exception as e:
            logging.error(f"{inspect.stack()[0][3]}: {e}")
            df.to_excel(filename, index=False)

        logging.debug(f"{df}")
        return True

    def query_form(self):
        '''
        busca de leis
        :return: return and write a json file
        '''
        self.queryid = self.firefox.window_handles
        self.firefox.get('http://www2.camara.leg.br/atividade-legislativa/legislacao/pesquisa/avancada')
        logging.debug(self.firefox.current_url)
        container = self.firefox.find_element_by_tag_name('body')
        if self.expressao:
            container = self.firefox.find_element_by_id('expressao')
            container.send_keys(self.expressao)
        if self.numero:
            container = self.firefox.find_element_by_id('numero')
            container.send_keys(self.numero)
        if self.ano:
            container = self.firefox.find_element_by_id('ano')
            container.send_keys(self.ano)

        if self.click:
            logging.debug(f"{['apelido', 'Ilei', 'Ileicom', 'Ideclei', 'Idecret']}")
            for item in self.click:
                self.firefox.find_element_by_id(item).click()
        # Desmarca apelido
        # self.firefox.find_element_by_id('apelido').click()
        # Lei ordinária
        # self.firefox.find_element_by_id('Ilei').click()
        # Lei complementar
        # self.firefox.find_element_by_id('Ileicom').click()
        # Decreto lei
        # self.firefox.find_element_by_id('Ideclei').click()
        # Decreto
        # self.firefox.find_element_by_id('Idecret').click()

        # É preciso passar o elemento para a classe
        origem = Select(self.firefox.find_element_by_name('origensF'))
        # Selecionar a opção pelo texto
        origem.select_by_visible_text('(Todas)')

        # É preciso passar o elemento para a classe
        situacao = Select(self.firefox.find_element_by_id('situacao'))
        # Selecionar a opção pelo valor
        situacao.select_by_value('')
        container.send_keys(Keys.ENTER)
        logging.debug('consultas selecionadas')

        logging.debug('title: {}'.format(self.firefox.title))
        logging.debug('url: '.format(self.firefox.current_url))
        logging.debug('id: {}'.format(self.queryid))

        # self.firefox.forward()
        logging.debug('confirmaçao id: {}'.format(self.firefox.window_handles))

        body = self.firefox.find_element_by_tag_name("body")
        body.send_keys(Keys.CONTROL + 't')

        try:
            wait = WebDriverWait(self.firefox, 10)
            wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Próxima")))
        except TimeoutException as e:
            logging.error(e)
            pass

        logging.debug(self.firefox.title)
        # assert "busca" in self.firefox.title
        atos = self.firefox.find_elements_by_class_name('titulo')
        self.atos2list(atos)
        while True:
            try:
                self.firefox.find_element_by_link_text('Próxima').click()
                self.atos2list(self.firefox.find_elements_by_class_name('titulo'))
            except Exception as e:
                logging.error(e)
                break

        logging.debug(f'collection_query: {len(self.collection_query)}')
        print('')
        for i in self.collection_query:
            logging.debug(f"{inspect.stack()[0][3]}{i}")

        assert "No results found." not in self.firefox.page_source
        # self.firefox.close()

    def fill_metadados(self, **kwargs):
        if 'jsonfile' in kwargs:
            df = pd.read_json(kwargs['jsonfile'])
        elif 'csvfile' in kwargs:
            df = pd.read_csv(kwargs['csvfile'])
        elif 'data' in kwargs:
            df = pd.DataFrame(kwargs['data'])
        else:
            df = pd.DataFrame(self.collection_query)
        print(df)
        meta_atos = []

        for title, link in zip(df['title'], df['link']):
            name = title.split(',')[0].split()[0].lower()
            try:
                numero = '{:0>5}{}'.format(*title.split(',')[0].split()[-1].split('-'))
            except Exception:
                numero = '{:0>5}'.format(title.split(',')[0].split()[-1])
            ano = title.split()[-1]
            url_referencias = link.strip()
            date = Legis.date_conv(title.split(',')[-1])
            try:
                conteudo = requests.get(url_referencias)
                logging.debug(conteudo.content)
                logging.debug(conteudo.status_code)
                if conteudo.status_code == 404:
                    raise URLError(conteudo.status_code)
            except URLError as e:
                print(e)

            soup = BeautifulSoup(conteudo.content, 'html.parser')
            try:
                container = soup.find(string=re.compile("Texto - Publicação Original")).parent.parent
                logging.debug(container)

                url_texto = urljoin(link, container['href'])
                print(url_texto)
            except Exception:
                url_texto = None
                logging.error(f"{inspect.stack()[0][3]}: {sys.exc_info()}")

            container = soup.find(string=re.compile('Origem')).parent.parent.find_next_sibling('span')
            logging.debug(container)

            origem = container.text
            print(origem)
            try:
                container = soup.find(string=re.compile('Situação: ')).parent.parent.find_next_sibling('span')
                logging.debug(container)
                situacao = container.text
                logging.debug(f'{situacao}')
            except Exception:
                situacao = None
                logging.error(f"{inspect.stack()[0][3]}: {sys.exc_info()}")

            container = soup.find(string=re.compile('Vide Norma'))
            logging.debug(container)

            normas_relacionadas = []
            if container:
                tags = container.parent.parent.find_next_sibling('ul').find_all('a')
                for tag in tags:
                    normas_relacionadas.append(re.sub('[\n\r\t]', ' ', tag.text).replace('  ', ''))
            logging.debug('NR: {normas_relacionadas}')

            container = soup.find(string=re.compile('Anexo'))
            print('>', container)

            anexos = []
            if container:
                try:
                    tags = container.parent.parent.find_next_sibling('ul').find_all('a')
                    print('>', tags)
                    for tag in tags:
                        anexos.append(urljoin(link, tag['href']))
                    print('>', anexos)
                except AttributeError:
                    pass
            self.collection_metadados_header = [
                'titulo', 'nome', 'numero', 'ano', 'data',
                'url_referencias', 'origem', 'situacao', 'normas_relacionadas', 'anexos', 'url_texto'
            ]
            meta_atos.append([
                title, name, numero, ano, date, url_referencias,
                origem, situacao, normas_relacionadas, anexos, url_texto
                ])
            logging.debug(meta_atos[-1])

            # logging.debug(len(atos))
        self.collection_metadados = meta_atos
        return self.collection_metadados

    @staticmethod
    def run():
        a = Scrapy_camara()
        a.ano = 1900
        # a.numero = 91
        a.filejson = 'selenium_multaccess03a.json'
        try:
            logging.debug('starting...')
            a.query_form()
            a.fill_metadados()
            # a.fill_content()
        except AttributeError as e:
            print(e)
            logging.error(e)
            raise
        finally:
            a.firefox.quit()
            pass


if __name__ == '__main__':
    # Scrapy_camara.run()
    # Scrapy_camara.wlogs()
    print(__file__)
    print(os.path.basename(__file__))
    pass
