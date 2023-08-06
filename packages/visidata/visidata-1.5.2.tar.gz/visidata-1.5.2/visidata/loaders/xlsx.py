
from visidata import *

def open_xlsx(p):
    vs = xlsxContents(p)
    return vs


class xlsxContents(Sheet):
    'Load XLSX file (in Excel Open XML format).'
    rowtype = 'sheets'  # rowdef: xlsxSheet
    columns = [
        Column('sheet', getter=lambda col,row: row.source.title),  # xlsx sheet title
        ColumnAttr('name', width=0),  # visidata Sheet name
        ColumnAttr('nRows', type=int),
        ColumnAttr('nCols', type=int),
    ]
    nKeys = 1

    def __init__(self, path):
        super().__init__(path.name, source=path)
        self.workbook = None

    @asyncthread
    def reload(self):
        import openpyxl
        self.workbook = openpyxl.load_workbook(self.source.resolve(), data_only=True, read_only=True)
        self.rows = []
        for sheetname in self.workbook.sheetnames:
            vs = xlsxSheet(joinSheetnames(self.name, sheetname), source=self.workbook[sheetname])
            vs.reload()
            self.rows.append(vs)

xlsxContents.addCommand(ENTER, 'dive-row', 'vd.push(cursorRow)')

class xlsxSheet(Sheet):
    @asyncthread
    def reload(self):
        worksheet = self.source

        self.columns = []
        self.rows = []

        rows = worksheet.iter_rows()
        hdrs = [list(wrapply(getattr, cell, 'value')
                   for cell in next(rows))
                       for i in range(options.header)
               ]
        colnames = ['\n'.join(str(hdr[i]) for i in range(len(hdr))) for hdr in zip(*hdrs)]
        for i, colname in enumerate(colnames):
            self.addColumn(ColumnItem(colname, i))

        for r in Progress(rows, total=worksheet.max_row or 0):
            row = list(wrapply(getattr, cell, 'value') for cell in r)
            for i in range(len(self.columns), len(row)):  # no-op if already done
                self.addColumn(ColumnItem(None, i, width=8))
            self.addRow(row)


class open_xls(Sheet):
    'Load XLS file (in Excel format).'
    rowtype = 'sheets'  # rowdef: xlsSheet
    columns = [
        Column('sheet', getter=lambda col,row: row.source.name),  # xls sheet name
        ColumnAttr('name', width=0),  # visidata sheet name
        ColumnAttr('nRows', type=int),
        ColumnAttr('nCols', type=int),
    ]
    nKeys = 1
    def __init__(self, path):
        super().__init__(path.name, source=path)
        self.workbook = None

    @asyncthread
    def reload(self):
        import xlrd
        self.workbook = xlrd.open_workbook(self.source.resolve())
        self.rows = []
        for sheetname in self.workbook.sheet_names():
            vs = xlsSheet(joinSheetnames(self.name, sheetname), source=self.workbook.sheet_by_name(sheetname))
            vs.reload()
            self.rows.append(vs)

open_xls.addCommand(ENTER, 'dive-row', 'vd.push(cursorRow)')

class xlsSheet(Sheet):
    @asyncthread
    def reload(self):
        worksheet = self.source
        self.columns = []
        if options.header:
            hdrs = [list(worksheet.cell(rownum, colnum).value for colnum in range(worksheet.ncols))
                        for rownum in range(options.header)]
            colnames = ['\\n'.join(str(hdr[i]) for i in range(len(hdr))) for hdr in zip(*hdrs)]
        else:
            colnames = ['']*worksheet.ncols

        for i, colname in enumerate(colnames):
            self.addColumn(ColumnItem(colname, i))

        self.rows = []
        for rownum in Progress(range(options.header, worksheet.nrows)):
            self.addRow(list(worksheet.cell(rownum, colnum).value for colnum in range(worksheet.ncols)))
