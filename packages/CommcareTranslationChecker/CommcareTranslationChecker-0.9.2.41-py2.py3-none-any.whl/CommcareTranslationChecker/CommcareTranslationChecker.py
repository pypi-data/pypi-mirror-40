from __future__ import print_function
import sys
import os
import datetime
import argparse
import traceback as tb
import openpyxl as xl

##### DEFINE GLOBALS #####
NON_LINGUISTIC_CHARACTERS = "~`!@#$%^&*()_-+={[}]|\\:;\"'<,>.?/"

##### DEFINE METHODS #####

def parseArguments():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="Location of Translation file to check", type=str, required = True)
    parser.add_argument("--columns", help="[Opt] Comma-separated list of column names to check. By default, all columns that start with 'default_' will be checked.", type=str, default=None)
    parser.add_argument("--base-column", help="[Opt] Name of column that others are to be compared against. Warnings are flagged for all columns that do not match the base-column. Defaults to leftmost column in columns.", type=str, default=None, dest='baseColumn')
    parser.add_argument("--ignore-order", help="[Opt] If passed, the order in which output value tags appear will not be considered when comparing cells against each other. This is useful if the order of the output value tags is different between columns because of differences in word orders between the languages involved.", action="store_true", default=False, dest='ignoreOrder')
    parser.add_argument("-v", "--verbose",  help="[Opt] If passed, output will be printed to the screen pointing out which rows of the file have issues.", action="store_true", default = False)
    parser.add_argument("--output-folder", help = "[Opt] Folder in which any output files should be passed. Defaults to 'commcareTranslationChecker_Output' folder relative to folder from which the script is called. Can be relative or absolute path.", type=str, default = "commcareTranslationChecker_Output", dest='outputFolder')
    parser.add_argument("--no-output-file", help = "[Opt] If passed, no output file will be created.", action="store_false", default = True, dest = "createOutputFileFlag")
    parser.add_argument("--configuration-sheet", help = "[Opt] Specify which sheet contains configuration information about modules and forms. Defaults to 'Modules_and_forms'", type=str, default = "Modules_and_forms", dest='configurationSheet')
    parser.add_argument("--configuration-sheet-column", help = "[Opt] specify which column in the configuration sheet contains expected sheet names. Defaults to 'sheet_name'", type=str, default = "sheet_name", dest='configurationSheetColumnName')
    parser.add_argument("--output-mismatch-types", help = "[Opt] If passed, information will be returned about the exact type of output value mismatch that occurs.", action="store_true", default = False, dest = "outputMismatchTypesFlag")
    parser.add_argument("--format-check", help = "[Opt] If passed, text formatting will be checked as well as output values.", action = "store_true", default = False, dest = "formatCheckFlag")
    parser.add_argument("--format-check-characters", help = "[Opt] A list of characters considered non-linguistic that will be counted when format-check is run. The characters \\ and \" need to be escaped as \\\\ and \\\". Defaults to CommcareTranslationChecker.NON_LINGUISTIC_CHARACTERS", type=str, default = None, dest = "formatCheckCharacters")
    parser.add_argument("--format-check-characters-add", help = "[Opt] A list of characters to be added to the default or passed format-check-characters list. The characters \\ and \" need to be escaped as \\\\ and \\\". Defaults to None.", type=str, default = None, dest = "formatCheckCharactersAdd")
    parser.add_argument("--debug-mode", "-d", action="store_true", default=False, dest="debugMode")
    return parser.parse_args()

def convertCellToOutputValueList(cell):
    '''
    Convert an Excel cell to a list of <output value...> tags contained within the cell

    Input:
    cell (xl.cell.cell.Cell): Cell whose contents are to be parsed

    Output:
    List of unicode objects, each representing an instance of <output value...> in cell. Any of these values that appear suspicious will be prepended with "ILL-FORMATTED TAG : "
    '''
    openTag = "<output value=\""
    closeTag ="\"/>"
    outputList = []
    currentIndex = 0
    try:
        while cell.value[currentIndex:].find(openTag) != -1:
            currentIndex += cell.value[currentIndex:].find(openTag) + len(openTag)
            closeTagIndex = cell.value[currentIndex:].find(closeTag)
            if closeTagIndex != -1:
                outputValue = cell.value[currentIndex:closeTagIndex + currentIndex]
                if outputValue.find(openTag) == -1:
                    outputList.append(outputValue)
                else:
                    outputList.append("ILL-FORMATTED TAG : " + outputValue)
            else:
                outputValue = cell.value[currentIndex:]
                print("closeTag not found for " + outputValue)
                outputList.append("ILL-FORMATTED TAG : " + outputValue)
    except TypeError as e:
        return []
    except Exception as e:
        print("FATAL ERROR determining output values for worksheet %s cell %s : %s" % (cell.parent.title, cell.coordinate, str(e)))
        exit(-1)

    return outputList

def createOutputCell(cell, wsOut):
    '''
    Make a copy of a Cell object into the exact same coordinates in the target Worksheet.

    Input:
    cell (xl.cell.cell.Cell): Cell whose contents and coordinates are to be copied
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet to which the cell's contents are to be copied

    Output:
    New Cell in wsOut
    '''
    try:
        newCell = wsOut.cell(coordinate = cell.coordinate) 
        newCell.value = cell.value
        newCell.style = xl.styles.Style(alignment = xl.styles.Alignment(wrap_text = True))
        return newCell
    except Exception as e:
        print("FATAL ERROR creating output cell for worksheet %s cell %s (writing to output worksheet %s) : %s" % (cell.parent.title, cell.coordinate, wsOut.title, str(e)))
        exit(-1)

def getOutputCell(cell, wsOut):
    '''
    Fetch an existing Cell object from wsOut corresponding to the coordinates of cell.

    Input:
    cell (xl.cell.cell.Cell): Cell whose coordinates are to be used to pinpoint target cell 
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet from which corresponding cell is to be pulled

    Output:
    Cell objects from wsOut corresponding to coordinates of cell 
    '''

    return wsOut[cell.coordinate]

def getNonLinguisticCharacterCount(val, additionalCharactersToCatch = None, characterList = None):
    '''
    Check a string for how many of each kind of non-linguistic character it contains and return a dictionary mapping character to count.

    Input:
    val(str): string to get counts from
    additionalCharactersToCatch(str [opt]): string of characters to append onto the characterList
    characterList(str [opt]): string of characters considered non-linguistic. Defaults to CommcareTranslationChecker.NON_LINGUISTIC_CHARACTERS

    Output:
    Dictionary mapping non-linguistic character to count of appearance in val 
    '''
    charCountDict = {}
    if val is None:
        val = ""

    if characterList == None:
        characterList = NON_LINGUISTIC_CHARACTERS
    if additionalCharactersToCatch != None:
        characterList += "".join([x for x in additionalCharactersToCatch if x not in characterList])

    for char in characterList:
        charCountDict[char] = val.count(char)

    return charCountDict


def checkRowForMismatch(row, columnDict, baseColumnIdx = None, ignoreOrder = False, wsOut = None, mismatchFlagIdx = None, outputMismatchTypesFlag = False, formatCheckFlag = False, formatCheckCharacters = None, formatCheckCharactersAdd = None):
    '''
    Check all of the given columns in a row provided for any mismatch in the columns' OutputValueList 

    Input:
    row(list): list of openyxl.cell.cell.Cell objects representing a single row in an Excel sheet 
    columnDict(dict): dictionary mapping column index to column name, representing every column to be checked against the baseColumn 
    baseColumnIdx(int [opt]): Index of the column to be considered "correct." Defaults to lowest-indexed column in columnDict.
    ignoreOrder(bool [opt]): If True, the order in which output values appear will be ignored for purposes of comparing cells. Otherwise, the order will matter. Defaults to False.
    wsOut(xl.worksheet.worksheet.Worksheet [opt]): Worksheet whose corresponding cell should be filled with Red if a mismatch occurs. Defaults to None.
    mismatchFlagIdx(int [opt]): Column index where the mismatchFlag value should be printed in wsOut
    outputMismatchTypesFlag(bool [opt]): Flag indicating whether to output the full mismatch types to the results file. Defaults to False
    formatCheckFlag(bool [opt]): Flag indicating whether to check for bad text formatting outside of output value. Defaults to False

    Output:
    Tuple consisting of a single-element dictionary mapping the baseColumn's index to its outputValueList, and a dictionary mapping the column indexes of mismatched cells to a tuple consisting of the associated cell's OutputValueList and a list of mismatchTypes. wsOut altered so that every Cell that is mismatched is filled with Red, and mismatchFlag column filled with "Y" if there was a mismatch in the row, "N" otherwise.
    '''
    mismatchDict = {}
    baseColumnDict=  {}
    baseFormatDict = {}

    baseOutputValueList = None

    mismatchFillStyle = xl.styles.Style(fill = xl.styles.PatternFill(fgColor = xl.styles.colors.Color(xl.styles.colors.RED), fill_type = "solid"), alignment = xl.styles.Alignment(wrap_text = True))
    lesserMismatchFillStyle = xl.styles.Style(fill = xl.styles.PatternFill(fgColor = xl.styles.colors.Color(xl.styles.colors.YELLOW), fill_type = "solid"), alignment = xl.styles.Alignment(wrap_text = True))

    ## Get columnDictKeyList for Python3
    columnDictKeyList = list(columnDict.keys())

    ## Build baseColumnDict
    if baseColumnIdx is None:
        baseColumnIdx = sorted(columnDictKeyList)[0]
    baseOutputValueList = convertCellToOutputValueList(row[baseColumnIdx])
    if ignoreOrder:
        baseOutputValueList = sorted(baseOutputValueList)
    baseColumnDict = {baseColumnIdx : baseOutputValueList}

    ## Build baseFormatDict if needed
    if formatCheckFlag:
        baseFormatDict = getNonLinguisticCharacterCount(row[baseColumnIdx].value, formatCheckCharacters, formatCheckCharactersAdd)

    for colIdx in columnDictKeyList:
        try:
            curOutputValueList = convertCellToOutputValueList(row[colIdx])
            if ignoreOrder:
                curOutputValueList = sorted(curOutputValueList)
            curFormatDict = {}
            if formatCheckFlag:
                curFormatDict = getNonLinguisticCharacterCount(row[colIdx].value, formatCheckCharacters, formatCheckCharactersAdd)
            if colIdx != baseColumnIdx and (baseOutputValueList != curOutputValueList or baseFormatDict != curFormatDict):
                ## Determine how everything is mismatched
                mismatchTypes = []

                ## Determine whether any ill-formatted tags exist:
                illFormattedValueList = []
                for value in curOutputValueList:
                    if value.startswith("ILL-FORMATTED TAG : "):
                        illFormattedValueList.append(value[20:])
                if illFormattedValueList != []:
                    mismatchTypes.append("Ill-Formatted Tags - " + ",".join(illFormattedValueList)) 

                ## Determine whether any values missing from current list
                missingValueList = []
                for value in baseOutputValueList:
                    if value not in curOutputValueList:
                        missingValueList.append(value)
                if missingValueList != []:
                    mismatchTypes.append("Missing Values - " + ",".join(missingValueList))

                ## Determine whether extra values have been added in current list
                extraValueList = []
                for value in curOutputValueList:
                    if value not in baseOutputValueList:
                        extraValueList.append(value)
                if extraValueList != []:
                    mismatchTypes.append("Extra Values - " + ",".join(extraValueList))

                ## Determine if, after considering missing/extra values, there are sort issues
                if not ignoreOrder and len(baseOutputValueList) != 0:
                    baseListIndex = 0
                    for value in curOutputValueList:
                        if value not in extraValueList:
                            while len(baseOutputValueList) > baseListIndex and baseOutputValueList[baseListIndex] in missingValueList:
                                baseListIndex += 1
                            if len(baseOutputValueList) > baseListIndex and value != baseOutputValueList[baseListIndex]:
                                mismatchTypes.append("Out of Order")
                                break 
                            baseListIndex += 1

                ## Determine whether there are any text formatting mismatches
                if baseFormatDict != curFormatDict:
                    formatDiffList = []
                    for key in baseFormatDict.keys():
                        keyDiff = curFormatDict[key] - baseFormatDict[key]
                        if keyDiff != 0:
                            formatDiffList.append("%s : %s" % (key, str(keyDiff) if keyDiff < 0 else "+" + str(keyDiff)))
                    mismatchTypes.append("Text Formatting Mismatch - " + ",".join(formatDiffList))

                if len(mismatchTypes) > 0:
                    mismatchDict[colIdx] = (curOutputValueList, mismatchTypes)

                if wsOut:
                    cellOut = getOutputCell(row[colIdx], wsOut)
                    if len(mismatchTypes) == 1 and "Text Formatting Mismatch" in mismatchTypes[0]:
                        curMismatchFillStyle = lesserMismatchFillStyle
                    else:
                        curMismatchFillStyle = mismatchFillStyle
                    cellOut.style = curMismatchFillStyle
                    if outputMismatchTypesFlag:
                        mismatchTypesColIdx = appendColumnIfNotExist(wsOut, "mismatch_%s"%(columnDict[colIdx],))
                        mismatchTypesCellOut = wsOut.rows[getOutputCell(row[0],wsOut).row-1][mismatchTypesColIdx]
                        #mismatchTypesCellOut = getOutputCell(row[0], wsOut).row[mismatchTypesColIdx]
                        mismatchTypesCellOut.value = ",".join(mismatchTypes)
                        mismatchTypesCellOut.style = curMismatchFillStyle

        except AttributeError as e:
            print(e)
        except Exception as e:
            print("FATAL ERROR comparing to baseColumn worksheet %s cell %s : %s" % (row[colIdx].parent.title, row[colIdx].coordinate, str(e)))
            tb.print_exc(e)
            exit(-1)

    mismatchCell =wsOut.cell(row = getOutputCell(row[0], wsOut).row, column = 1).offset(column = mismatchFlagIdx)
    if len(mismatchDict) > 0:
        curMismatchFillStyle = lesserMismatchFillStyle
        for key in mismatchDict:
            if len(mismatchDict[key][1]) > 1 or "Text Formatting Mismatch" not in mismatchDict[key][1][0]:
                curMismatchFillStyle = mismatchFillStyle
        mismatchCell.value = "Y"
        mismatchCell.style = curMismatchFillStyle
    else:
        mismatchCell.value = "N"

    return (baseColumnDict, mismatchDict)

def appendColumnIfNotExist(ws, columnHeader):
    '''
    Check whether a column with the given header already exists in ws, and append it if not.

    Input:
    ws (xl.worksheet.worksheet.worksheet): Worksheet to append column to
    columnHeader (str): Header of new column

    Output:
    If column with columnHeader does not exist, append it to ws. Return index of column with columnHeader.
    '''
    maxHeaderIdx = 0
    for headerIdx, cell in enumerate(ws.rows[0]):
        if cell.value == columnHeader:
            return headerIdx
        maxHeaderIdx = max(headerIdx, maxHeaderIdx)
    newColIdx = maxHeaderIdx + 1
    ws.cell("A1").offset(column = newColIdx).value = columnHeader
    return newColIdx



def checkConfigurationSheet(wb, ws, configurationSheetColumnName, wsOut, verbose = False):
    '''
    Check that the workbook contains one sheet for every corresponding entry in the configurationSheetColumn of ws, and highlight all cells in wsOut that represent sheets that don't exist.

    Input:
    wb (xl.workbook.workbook.Workbook): Workbook containing sheets to check against configurationSheetColumn
    ws (xl.worksheet.worksheet.Worksheet): Worksheet containing the column to check 
    configurationSheetColumnName (str): Name of column to compare sheet names against
    wsOut (xl.worksheet.worksheet.Worksheet): Worksheet to print highlighted cells to 
    verbose (boolen [opt]): If passed, prints each missing sheet to the screen

    Output:
    List of sheets that are missing from the Workbook. If configurationSheetColumnName does not exist in ws, returns None
    '''

    mismatchFillStyle = xl.styles.Style(fill = xl.styles.PatternFill(fgColor = xl.styles.colors.Color(xl.styles.colors.RED), fill_type = "solid"), alignment = xl.styles.Alignment(wrap_text = True))
    missingSheetList = []

    ## Check that the configuration column exists at all
    colIdx = None
    for headerIdx, cell in enumerate(ws.rows[0]):
        if cell.value == configurationSheetColumnName:
            colIdx = headerIdx
    if colIdx == None:
        print(configurationSheetColumnName, " not found in ", ws.title, ". Skipping sheet check.")
        return None

    ## Iterate over configuration column, flagging red if corresponding sheet does not exist
    for cell in ws.columns[colIdx][1:]:
        if cell.value not in (sheet.title for sheet in wb):
            missingSheetList.append(cell.value)
            getOutputCell(cell, wsOut).style = mismatchFillStyle
            if verbose:
                print("WARNING: This sheet is missing from the workbook: %s" % (cell.value,))

    return missingSheetList


def main(argv):
    args = parseArguments()
    try:
        wb = xl.load_workbook(args.file)
        if args.verbose:
            print("Workbook Loaded")
    except xl.exceptions.InvalidFileException as e:
        print("Invalid File: %s" % (str(e),))
        if args.debugMode:
            tb.print_exc(e)
        exit(-1)

    ## Open new Workbook
    wbOut = xl.Workbook()
    wbOut.remove_sheet(wbOut.active)

    ## Summary lists
    wsMismatchDict = {}
    wbMissingSheets = []

    ## Iterate through WorkSheets
    for ws in wb:
        try:
            wbOut.create_sheet(title = ws.title)
            wsOut = wbOut[ws.title]


            ## Dictionaries mapping column index to column name
            defaultColumnDict = {}
            mismatchTypesColumnDict = {}

            maxHeaderIdx = 0
            ## Find all columns of format "default_[CODE]"
            for headerIdx, cell in enumerate(ws.rows[0]):
                ## First, copy cell into new workbook
                cellOut = createOutputCell(cell, wsOut)
                if args.columns:
                    if cell.value in args.columns:
                        defaultColumnDict[headerIdx] = cell.value
                elif cell.value and cell.value[:8] == "default_":
                    defaultColumnDict[headerIdx] = cell.value
                if headerIdx > maxHeaderIdx:
                    maxHeaderIdx = headerIdx
            ## If defaultColumnDict is empty, skip processing
            ## Otherwise, create header cell in wsOut for mismatchFlag
            if len(defaultColumnDict) != 0:
                mismatchFlagIdx = appendColumnIfNotExist(wsOut, "mismatchFlag")


                for rowIdx, row in enumerate(ws.rows[1:]):
                    ## First, copy every cell into new workbook
                    for cell in row:
                        cellOut = createOutputCell(cell, wsOut)

                    ## Fetch baseColumn information
                    baseColumnIdx = None
                    if args.baseColumn:
                        for colIdx in defaultColumnDict.keys():
                            if defaultColumnDict[colIdx] == args.baseColumn:
                                baseColumnIdx = colIdx 

                    ## Check row for mismatch and print results
                    rowCheckResults = checkRowForMismatch(row, defaultColumnDict, baseColumnIdx, args.ignoreOrder, wsOut, mismatchFlagIdx, args.outputMismatchTypesFlag, args.formatCheckFlag, args.formatCheckCharactersAdd, args.formatCheckCharacters)
                    if len(rowCheckResults[1]) > 0:
                        if ws.title not in wsMismatchDict.keys():
                            wsMismatchDict[ws.title] = 1
                        else:
                            wsMismatchDict[ws.title] += 1
                        if args.verbose:
                            baseColumnName = defaultColumnDict[list(rowCheckResults[0].keys())[0]]
                            if args.outputMismatchTypesFlag:
                                mismatchColumnNames = ",".join("%s (%s)" % (defaultColumnDict[i], ",".join(rowCheckResults[1][i][1])) for i in rowCheckResults[1].keys())
                            else:
                                mismatchColumnNames = ",".join(defaultColumnDict[i] for i in rowCheckResults[1].keys())
                            print("WARNING %s row %s: the output values in %s do not match %s" % (ws.title, rowIdx+2, mismatchColumnNames, baseColumnName))
            elif args.verbose:
                print("WARNING %s: No columns found for comparison" % (ws.title,))

            ## If ws is a configuration sheet, run the configuration check
            if ws.title == args.configurationSheet:
                wbMissingSheets = checkConfigurationSheet(wb, ws, args.configurationSheetColumnName, wsOut, args.verbose)
        except Exception as e:
            print("FATAL ERROR in worksheet %s : %s" % (ws.title, str(e)))
            if args.debugMode:
                tb.print_exc(e)
            exit(-1)

    ## Save workbook and print summary
    if len(wsMismatchDict) > 0 or (wbMissingSheets is not None and len(wbMissingSheets) > 0):
        if args.createOutputFileFlag:
            tsString = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            fileBasename = os.path.splitext(os.path.basename(args.file))[0]
            outputFolder = args.outputFolder
            outputFileName = os.path.join(outputFolder,"%s_%s_Output.xlsx" % (fileBasename, tsString))
            ## Create the output directory if it does not exist
            if not os.path.exists(os.path.dirname(outputFileName)):
                try:
                    os.makedirs(os.path.dirname(outputFileName),)
                    print("Output directory did not exist, created %s" % (os.path.dirname(outputFileName),))
                except OSError as e:
                    if e.errorno != e.EEXIST:
                        print("ERROR CREATING OUTPUT DIRECTORY : %s" % (str(e),))
                        if args.debugMode:
                            tb.print_exc(e)
            wbOut.save(outputFileName)
            print("There were issues with the following worksheets, see %s for details:" % (outputFileName,))
        else:
            print("There were issues with the following worksheets:")
        if wbMissingSheets is not None:
            for sheet in wbMissingSheets:
                print("%s is missing from the workbook." % (sheet,))
        for key in wsMismatchDict.keys():
            print("%s : %s row%s mismatched" % (key, wsMismatchDict[key], "" if wsMismatchDict[key]==1 else "s"))

def entryPoint():
    main(sys.argv[1:])

if __name__ == "__main__":
    entryPoint()